import os
import datetime
from werkzeug import secure_filename
from flask_login import current_user
from app import db, telomere
from app.model.manifest import Manifest
from app.model.samplePlate import SamplePlate
from openpyxl import load_workbook
import traceback
from decimal import *
from itertools import izip_longest


class ManifestService():

    def SaveAndReturn(self, manifestFile):
        filename = secure_filename(manifestFile.filename)

        manifest = Manifest(
            filename=filename,
            uploaded=datetime.datetime.now(),
            userId=current_user.id
        )

        db.session.add(manifest)
        db.session.flush()

        manifestFile.save(self.GetPath(manifest))

        return manifest

    def Process(self, manifest):
        wb = load_workbook(filename=self.GetPath(manifest), use_iterators=True)
        ws = wb.worksheets[0]

        try:
            # Divide the spreadsheet rows into chunks of
            # no more than 5000 rows.  This is because the
            # UoL LAMP servers are running python 2.6
            # which for some reason errors if you
            # try to insert too many rows at a time.
            # I think it's because the query text gets
            # too big.
            chunks = izip_longest(
                *[iter(ws.iter_rows(row_offset=1))] * 5000,
                fillvalue=None
            )

            for c in chunks:
                db.session.execute(
                    # Works only for MySQL.  Ignore duplicates
                    'INSERT IGNORE INTO sample (sampleCode) VALUES (:sampleCode)',
                    [{"sampleCode": row[2].value} for row in c if row is not None]
                )
                db.session.flush()

            # Have to recreate the chunks because they've
            # been consumed.  They're so meaty!
            chunks = izip_longest(
                *[iter(ws.iter_rows(row_offset=1))] * 5000,
                fillvalue=None
            )

            for c in chunks:
                db.session.bulk_insert_mappings(
                    SamplePlate,
                    [{"plateName": row[0].value,
                        "well": row[1].value,
                        "sampleCode": row[2].value,
                        "conditionDescription": row[3].value,
                        "volume": row[4].value,
                        "dnaTest": self._roundDecimal(row[5].value),
                        "picoTest": self._roundDecimal(row[6].value),
                        "manifestId": manifest.id}
                        for row in c if row is not None]
                )

                db.session.flush()

        except:
            telomere.logger.error(traceback.format_exc())
            return False

        return True

    def _roundDecimal(self, value):
        if value is None:
            return None

        return Decimal(
            Decimal(str(value)).quantize(
                Decimal('.01'),
                rounding=ROUND_HALF_UP)
        )

    def ValidateFormat(self, manifest):

        wb = load_workbook(filename=self.GetPath(manifest), use_iterators=True)
        ws = wb.worksheets[0]

        return (
            ws['A1'].value == 'Plate Name' and
            ws['B1'].value == 'WELL' and
            ws['C1'].value == 'Sample_Name' and
            ws['D1'].value == 'Condition' and
            ws['E1'].value == 'Volume' and
            ws['F1'].value == 'DNA Test (ng/ul)' and
            ws['G1'].value == 'Pico Test (ng/ul)')

    def GetPath(self, manifest):
        return os.path.join(
            telomere.config['SPREADSHEET_UPLOAD_DIRECTORY'],
            self.GetFilename(manifest)
        )

    def GetFilename(self, manifest):
        return "manifest_%d.xlsx" % manifest.id
