import os
import datetime
import re
from werkzeug import secure_filename
from flask_login import current_user
from flask import flash
from app import db, telomere
from app.services.batch import BatchService
from app.services.outstandingError import OutstandingErrorService
from app.model.spreadsheet import Spreadsheet
from app.model.measurement import Measurement
from app.model.sample import Sample
from openpyxl import load_workbook


class SpreadsheetService():

    def SaveAndReturn(self, spreadsheetFile, batch):
        filename = secure_filename(spreadsheetFile.filename)

        spreadsheet = Spreadsheet(
            filename=filename,
            uploaded=datetime.datetime.now(),
            userId=current_user.id,
            batchId=batch.id
        )

        db.session.add(spreadsheet)
        db.session.flush()

        spreadsheetFile.save(self.GetPath(spreadsheet))

        return spreadsheet

    def Process(self, spreadsheet, disallowPlateNameMismatch):
        result = SpreadsheetLoadResult()

        outstandingErrorService = OutstandingErrorService()

        wb = load_workbook(filename=self.GetPath(spreadsheet), read_only=True)
        ws = wb.worksheets[0]

        for row in ws.iter_rows(row_offset=1):
            sampleCode = row[23].value  # Col X

            if (sampleCode is None or
                    not (
                        str(sampleCode).isdigit() or
                        sampleCode == Sample.POOL_NAME)):
                continue

            sample = Sample.query.filter(
                Sample.sampleCode == sampleCode).first()

            if sample is None:
                result.missingSampleCodes.add(sampleCode)
                continue

            if (sample.plate_name_mismatch(spreadsheet.batch.plateName) and
                    disallowPlateNameMismatch):

                result.incorrectPlateName.add(sampleCode)
                continue

            measurement = Measurement(
                batchId=spreadsheet.batch.id,
                sampleId=sample.id,
                t_to=row[1].value,  # Col B
                t_amp=row[2].value,  # Col C
                t=row[3].value,  # Col D
                s_to=row[13].value,  # Col N
                s_amp=row[14].value,  # Col O
                s=row[15].value,  # Col P
                primerBatch=spreadsheet.batch.primerBatch,
                errorCode=row[29].value or ''  # Col AD
            )

            for oe in sample.outstandingErrors:
                outstandingErrorService.CompleteError(oe)

            db.session.add(measurement)

        batchService = BatchService()
        batchService.SetCoefficientsOfVariation(spreadsheet.batch)

        db.session.flush()

        for e in batchService.GetValidationErrors(spreadsheet.batch):
            db.session.add(e)
            db.session.flush()

            if e.sample.has_good_measurement():
                outstandingErrorService.CompleteError(e)
            else:
                result.hasOutstandingErrors = True

        return result

    def ValidateFormat(self, spreadsheet):

        wb = load_workbook(
            filename=self.GetPath(spreadsheet),
            read_only=True)

        ws = wb.worksheets[0]

        return (
            ws['A1'].value == 'Name' and
            ws['B1'].value == 'Take Off' and
            ws['C1'].value == 'Amplification' and

            ws['D1'].value == 'Comparative Conc.' and
            ws['E1'].value == 'Rep. Takeoff' and
            ws['F1'].value == 'Rep. Takeoff (95% CI)' and

            ws['G1'].value == 'Rep. Amp.' and
            ws['H1'].value == 'Rep. Amp. (95% CI)' and
            ws['I1'].value == 'Rep. Conc.' and
            ws['J1'].value == 'Rep. Calibrator' and

            ws['M1'].value == 'Name' and
            ws['N1'].value == 'Take Off' and
            ws['O1'].value == 'Amplification' and
            ws['P1'].value == 'Comparative Conc.' and
            ws['Q1'].value == 'Rep. Takeoff' and
            ws['R1'].value == 'Rep. Takeoff (95% CI)' and
            ws['S1'].value == 'Rep. Amp.' and
            ws['T1'].value == 'Rep. Amp. (95% CI)' and
            ws['U1'].value == 'Rep. Conc.' and
            ws['V1'].value == 'Rep. Calibrator' and

            ws['X1'].value == 'id' and
            ws['Y1'].value == 't' and
            ws['Z1'].value == 's' and
            ws['AA1'].value == 'ts' and
            ws['AB1'].value == 'ave ts' and
            ws['AC1'].value == 'cv' and
            ws['AD1'].value == 'error code'
        )

    def GetPath(self, spreadsheet):
        return os.path.join(
            telomere.config['SPREADSHEET_UPLOAD_DIRECTORY'],
            self.GetFilename(spreadsheet)
        )

    def GetFilename(self, spreadsheet):
        return "%d.xlsx" % spreadsheet.id

    def _isValidValue(self, value):
        valAsString = str(value)
        p = re.compile('\d+(\.\d+)?')
        return p.match(valAsString) is not None


class SpreadsheetLoadResult:

    def __init__(self, *args, **kwargs):
        self.missingSampleCodes = Set()
        self.incorrectPlateName = Set()
        self.hasOutstandingErrors = False

    def abortUpload(self):
        return (
            len(self.missingSampleCodes) > 0 or
            len(self.incorrectPlateName) > 0
        )
